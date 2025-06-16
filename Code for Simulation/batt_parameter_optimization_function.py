"""
Battery Parameter Estimation using Bayesian Optimization

This module implements battery parameter estimation using Bayesian optimization
for different battery working conditions and batches.
"""

import warnings
import os
from collections import abc
from numbers import Number
from typing import List, Tuple, Optional, Dict, Any

import numpy as np
import scipy.io
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d

# Third-party libraries
from bayes_opt import BayesianOptimization, UtilityFunction
from openpyxl import Workbook, load_workbook
from progpy.loading import Piecewise
from progpy.state_estimators import UnscentedKalmanFilter
from progpy.utils.containers import InputContainer, OutputContainer

# Suppress warnings
warnings.filterwarnings('ignore')


class BatteryParameterEstimator:
    """
    A class for estimating battery parameters using Bayesian optimization.
    """

    # Battery working condition configurations
    BATCH_CONFIGS = {
        'batch01': {'Mid_SOC': 0.8, 'DOD': 0.2},
        'batch02': {'Mid_SOC': 0.7, 'DOD': 0.2},
        'batch03': {'Mid_SOC': 0.5, 'DOD': 0.2},
        'batch04': {'Mid_SOC': 0.4, 'DOD': 0.2},
        'batch05': {'Mid_SOC': 0.3, 'DOD': 0.2},
        'batch06': {'Mid_SOC': 0.2, 'DOD': 0.2},
        'batch07': {'Mid_SOC': 0.75, 'DOD': 0.3},
        'batch08': {'Mid_SOC': 0.6, 'DOD': 0.3},
        'batch09': {'Mid_SOC': 0.5, 'DOD': 0.3},
        'batch10': {'Mid_SOC': 0.45, 'DOD': 0.3},
        'batch11': {'Mid_SOC': 0.7, 'DOD': 0.4},
        'batch12': {'Mid_SOC': 0.55, 'DOD': 0.4},
        'batch13': {'Mid_SOC': 0.5, 'DOD': 0.4},
        'batch14': {'Mid_SOC': 0.4, 'DOD': 0.4},
        'batch15': {'Mid_SOC': 0.3, 'DOD': 0.4},
        'batch16': {'Mid_SOC': 0.5, 'DOD': 0.5},
        'batch17': {'Mid_SOC': 0.35, 'DOD': 0.5},
        'batch18': {'Mid_SOC': 0.5, 'DOD': 0.6},
        'batch19': {'Mid_SOC': 0.5, 'DOD': 0.7},
        'batch20': {'Mid_SOC': 0.5, 'DOD': 0.8},
        'batch21': {'Mid_SOC': 0.5, 'DOD': 1.0}
    }

    # Default simulation parameters
    DEFAULT_VOLTAGE_BOUNDS = (2.7, 4.5)
    DEFAULT_CURRENT = 5.2
    DEFAULT_TEMP_OFFSET = 274.15
    DEFAULT_VEOD = 2.75
    DEFAULT_SOC_SAMPLES = 100

    @staticmethod
    def get_battery_working_condition(batch_name: str) -> Tuple[float, float]:
        """
        Get battery working conditions (Mid_SOC, DOD) for a given batch.

        Args:
            batch_name: Name of the battery batch

        Returns:
            Tuple of (Mid_SOC, DOD) values

        Raises:
            ValueError: If batch_name is not found in configurations
        """
        if batch_name not in BatteryParameterEstimator.BATCH_CONFIGS:
            raise ValueError(f"Unknown batch name: {batch_name}")

        config = BatteryParameterEstimator.BATCH_CONFIGS[batch_name]
        return config['Mid_SOC'], config['DOD']

    @staticmethod
    def rescale_array(arr: np.ndarray, new_min: float = 0.7, new_max: float = 0.9) -> np.ndarray:
        """
        Rescale array values to a new range.

        Args:
            arr: Input array to rescale
            new_min: New minimum value
            new_max: New maximum value

        Returns:
            Rescaled array
        """
        arr_min, arr_max = arr.min(), arr.max()
        if arr_max == arr_min:
            return np.full_like(arr, (new_min + new_max) / 2)
        return ((arr - arr_min) / (arr_max - arr_min)) * (new_max - new_min) + new_min

    @staticmethod
    def create_folder(folder_path: str) -> None:
        """Create folder if it doesn't exist."""
        os.makedirs(folder_path, exist_ok=True)

    @staticmethod
    def write_excel_parameters(excel_path: str, qMax: float, R0: float, wr: float) -> None:
        """
        Write important parameters to Excel file.

        Args:
            excel_path: Path to Excel file
            qMax: Maximum charge parameter
            R0: Internal resistance parameter
            wr: Aging parameter
        """
        if not os.path.exists(excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['qMax', 'R0', 'wr'])  # Header row
        else:
            workbook = load_workbook(excel_path)
            sheet = workbook.active

        sheet.append([qMax, R0, wr])
        workbook.save(excel_path)

    @staticmethod
    def write_excel_results(excel_path: str, values: List[float]) -> None:
        """
        Write simulation results to Excel file.

        Args:
            excel_path: Path to Excel file
            values: List of result values
        """
        headers = ['t', 'v', 'tb', 'Vo', 'Vsn', 'Vsp', 'qnB', 'qnS', 'qpB', 'qpS', 'qMax', 'Ro', 'D']

        if not os.path.exists(excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
        else:
            workbook = load_workbook(excel_path)
            sheet = workbook.active

        sheet.append(values)
        workbook.save(excel_path)


def validate_inputs(keys: List[str], batt, times, inputs, outputs) -> None:
    """
    Validate input parameters for parameter estimation.

    Args:
        keys: List of parameter keys to optimize
        batt: Battery model object
        times: Time data
        inputs: Input data
        outputs: Output data

    Raises:
        ValueError: If validation fails
        TypeError: If input types are incorrect
    """
    # Validate keys
    if isinstance(keys, set):
        raise ValueError("Keys cannot be a Set. Sets are unordered by construction.")

    for key in keys:
        if key not in batt.parameters:
            raise ValueError(f"Key '{key}' not in model parameters")

    # Validate data types
    if isinstance(times, set) or isinstance(inputs, set) or isinstance(outputs, set):
        raise TypeError("Times, inputs, and outputs cannot be sets.")

    # Validate data lengths
    if len(times) != len(inputs) or len(inputs) != len(outputs):
        raise ValueError(
            f"Times, inputs, and outputs must be same length. "
            f"Got lengths: times={len(times)}, inputs={len(inputs)}, outputs={len(outputs)}"
        )

    if len(times) == 0:
        raise ValueError("Times, inputs, and outputs must have at least one element")


def prepare_bounds(bounds, keys: List[str], batt) -> List[Tuple[float, float]]:
    """
    Prepare optimization bounds for parameters.

    Args:
        bounds: Bounds specification (dict or list of tuples)
        keys: List of parameter keys
        batt: Battery model object

    Returns:
        List of (lower, upper) bound tuples

    Raises:
        ValueError: If bounds specification is invalid
    """
    if isinstance(bounds, dict):
        # Validate dict keys
        for key in bounds.keys():
            if key not in batt.parameters:
                warnings.warn(f"{key} is not a valid parameter")

        return [bounds.get(key, (-np.inf, np.inf)) for key in keys]

    if not isinstance(bounds, abc.Iterable):
        raise ValueError(f"Bounds must be a tuple of tuples or a dict, got {type(bounds)}")

    if len(bounds) != len(keys):
        raise ValueError(
            f"Bounds length ({len(bounds)}) must match keys length ({len(keys)}). "
            "Use a dict for partial bounds definition."
        )

    # Validate individual bounds
    for i, bound in enumerate(bounds):
        if isinstance(bound, set):
            raise TypeError(f"Bound {bound} cannot be a Set")
        if not isinstance(bound, abc.Iterable) or len(bound) != 2:
            raise ValueError(f"Each bound must be a tuple (lower, upper), got {type(bound)}")

    return bounds


def simulate_battery_discharge(batt, runs, optimization_params: Dict[str, float]) -> Tuple[np.ndarray, np.ndarray]:
    """
    Simulate battery discharge with given parameters.

    Args:
        batt: Battery model object
        runs: Simulation run data
        optimization_params: Dictionary of optimization parameters (qMax, Ro, wr)

    Returns:
        Tuple of (simulated_voltage, simulated_temperature) arrays
    """
    # Set optimization parameters
    batt.parameters.data['wr'] = optimization_params['wr']
    batt.parameters['Ro'] = optimization_params['Ro']
    batt.parameters['qMax'] = optimization_params['qMax']

    # Update initial state parameters
    batt.parameters['x0']['qMax'] = optimization_params['qMax']
    batt.parameters['x0']['Ro'] = optimization_params['Ro']
    batt.parameters['x0']['tb'] = runs[0][0][0] + BatteryParameterEstimator.DEFAULT_TEMP_OFFSET
    batt.parameters['tb'] = runs[0][0][0] + BatteryParameterEstimator.DEFAULT_TEMP_OFFSET
    batt.parameters['VEOD'] = BatteryParameterEstimator.DEFAULT_VEOD

    # Disable noise for simulation
    batt.parameters['measurement_noise'] = 0
    batt.parameters['process_noise'] = 0

    # Adjust charge parameters
    Q_transfer_parameter = optimization_params['qMax'] / (batt.parameters['qMaxThreshold'] / 0.7)
    charge_params = ['qnS', 'qnB', 'qpS', 'qpB']
    for param in charge_params:
        batt.parameters['x0'][param] *= Q_transfer_parameter

    # Perform state estimation (using batch21 data as reference)
    data_all = scipy.io.loadmat('')['RetiredBatteryData_all'][0, 0]  # Path needs to be provided
    ref_time = data_all['batch21'][0, 1][0][0][0][13][0, 0]
    ref_voltage = data_all['batch21'][0, 1][0][0][0][7][0, 0]
    ref_temp = data_all['batch21'][0, 1][0][0][0][10][0, 0]

    filt = UnscentedKalmanFilter(batt, batt.initialize())
    filt.estimate(ref_time, {'i': BatteryParameterEstimator.DEFAULT_CURRENT},
                  {'t': ref_temp, 'v': ref_voltage})
    batt.parameters['x0'] = filt.x.mean

    # Simulate discharge
    options = {'print': False, 'progress': False, 'dt': 2}
    future_loading = Piecewise(batt.InputContainer, [float('inf')],
                               {'i': [BatteryParameterEstimator.DEFAULT_CURRENT]})

    simulated_results = batt.simulate_to_threshold(future_loading, **options)

    # Extract voltage and temperature data within valid range
    simulated_voltage = []
    simulated_temperature = []

    for output in simulated_results.outputs.data:
        v_min, v_max = BatteryParameterEstimator.DEFAULT_VOLTAGE_BOUNDS
        if v_min < output['v'] < v_max:
            simulated_voltage.append(output['v'])
            simulated_temperature.append(output['t'])

    return np.array(simulated_voltage), np.array(simulated_temperature)


def calculate_optimization_error(simulated_voltage: np.ndarray, true_voltage: List[float],
                                 Mid_SOC: float, DOD: float) -> float:
    """
    Calculate optimization error between simulated and true voltage data.

    Args:
        simulated_voltage: Simulated voltage array
        true_voltage: True voltage measurements
        Mid_SOC: Middle state of charge
        DOD: Depth of discharge

    Returns:
        Negative error value (for maximization)
    """
    start_SOC = Mid_SOC + DOD / 2
    end_SOC = Mid_SOC - DOD / 2

    # Create SOC alignment using rescaling
    sample_SOC = BatteryParameterEstimator.DEFAULT_SOC_SAMPLES
    SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(sample_SOC) / (sample_SOC - 1), new_min=end_SOC, new_max=start_SOC
    )

    # Interpolate observed voltage
    observed_SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(len(true_voltage)) / (len(true_voltage) - 1),
        new_min=end_SOC, new_max=start_SOC
    )
    func_observed = interp1d(np.flipud(observed_SOC_range), true_voltage, kind='cubic')
    observed_voltage_range = func_observed(SOC_range)

    # Interpolate simulated voltage
    simulated_SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(len(simulated_voltage)) / (len(simulated_voltage) - 1),
        new_min=0, new_max=1
    )
    func_simulated = interp1d(np.flipud(simulated_SOC_range), simulated_voltage, kind='cubic')
    simulated_voltage_range = func_simulated(SOC_range)

    # Calculate weighted error (penalize positive errors more)
    error_diff = simulated_voltage_range - observed_voltage_range
    weighted_error = np.where(error_diff > 0, 2 * error_diff ** 2, error_diff ** 2)

    return -np.sum(weighted_error)


def estimate_params(
        batt,
        battery_initial_state,
        Mid_SOC: float,
        DOD: float,
        batch_name: str,
        battery_number: int,
        cycle_number: int,
        battery_data_path: str,
        saving_file_path: str,
        optimization_iter: int = 100,
        runs: Optional[List[Tuple]] = None,
        keys: Optional[List[str]] = None,
        times: Optional[List[float]] = None,
        inputs: Optional[List[InputContainer]] = None,
        outputs: Optional[List[OutputContainer]] = None,
        method: str = 'nelder-mead',
        **kwargs
) -> Tuple[float, float, float]:
    """
    Estimate battery parameters using Bayesian optimization.

    Args:
        batt: Battery model object
        battery_initial_state: Initial state of the battery
        Mid_SOC: Middle state of charge
        DOD: Depth of discharge
        batch_name: Name of the battery batch
        battery_number: Battery number identifier
        cycle_number: Cycle number identifier
        battery_data_path: Path to battery data file
        saving_file_path: Path to save results
        optimization_iter: Number of optimization iterations
        runs: List of (times, inputs, outputs) tuples
        keys: List of parameter keys to optimize
        times: Time data
        inputs: Input data
        outputs: Output data
        method: Optimization method (legacy parameter)
        **kwargs: Additional configuration options

    Returns:
        Tuple of optimized (qMax, Ro, wr) parameters
    """
    # Load battery data
    data_all = scipy.io.loadmat(battery_data_path)['RetiredBatteryData_all'][0, 0]

    # Set default keys if not provided
    if keys is None:
        keys = [key for key in batt.parameters.keys()
                if isinstance(batt.parameters[key], Number)]

    # Validate inputs
    validate_inputs(keys, batt, times or [], inputs or [], outputs or [])

    # Prepare configuration
    config = {
        'error_method': 'MSE',
        'bounds': tuple((-np.inf, np.inf) for _ in keys),
        'options': None,
        'tol': 1e-6
    }
    config.update(kwargs)

    # Prepare data for optimization
    if runs is None and all(x is not None for x in [times, inputs, outputs]):
        # Convert to list format if needed
        for data in [times, inputs, outputs]:
            if isinstance(data, np.ndarray):
                data = data.tolist()

        # Ensure proper list structure
        if not isinstance(times[0], (abc.Sequence, np.ndarray)):
            times = [times]
        if not isinstance(inputs[0], (abc.Sequence, np.ndarray)):
            inputs = [inputs]
        if not isinstance(outputs[0], (abc.Sequence, np.ndarray)):
            outputs = [outputs]

        runs = list(zip(times, inputs, outputs))

    if runs is None:
        raise ValueError("Either 'runs' or 'times', 'inputs', 'outputs' must be provided")

    # Prepare bounds
    bounds = prepare_bounds(config['bounds'], keys, batt)

    # Convert container types if needed
    for i, (run_times, run_inputs, run_outputs) in enumerate(runs):
        has_changed = False

        if not isinstance(run_inputs[0], batt.InputContainer):
            run_inputs = [batt.InputContainer(u) for u in run_inputs]
            has_changed = True

        if isinstance(run_outputs, np.ndarray):
            run_outputs = [batt.OutputContainer(u) for u in run_outputs]
            has_changed = True

        if has_changed:
            runs[i] = (run_times, run_inputs, run_outputs)

    # Extract true voltage data
    true_voltage = [x['v'] for x in runs[0][2]]

    def optimization_function(Ro: float, qMax: float, wr: float) -> float:
        """
        Objective function for Bayesian optimization.

        Args:
            Ro: Internal resistance parameter
            qMax: Maximum charge parameter
            wr: Aging parameter

        Returns:
            Negative error value (for maximization)
        """
        try:
            # Set initial state
            batt.parameters['x0'] = battery_initial_state

            # Simulate with current parameters
            optimization_params = {'Ro': Ro, 'qMax': qMax, 'wr': wr}
            simulated_voltage, _ = simulate_battery_discharge(batt, runs, optimization_params)

            # Calculate and return error
            return calculate_optimization_error(simulated_voltage, true_voltage, Mid_SOC, DOD)

        except Exception as e:
            print(f"Error in optimization function: {e}")
            return -1e6  # Return large negative value for failed simulations

    # Setup Bayesian optimization
    param_bounds = {'qMax': bounds[0], 'Ro': bounds[1], 'wr': bounds[2]}
    bo = BayesianOptimization(f=optimization_function, pbounds=param_bounds)

    # Configure acquisition function
    acquisition_function = UtilityFunction(kind="poi", xi=0.3e-4)

    # Run optimization
    bo.maximize(n_iter=optimization_iter, acquisition_function=acquisition_function)

    # Extract optimal parameters
    optimal_params = bo.max['params']
    qMax_opt = optimal_params['qMax']
    Ro_opt = optimal_params['Ro']
    wr_opt = optimal_params['wr']

    # Generate final simulation with optimal parameters
    batt.parameters['x0'] = battery_initial_state
    final_sim_voltage, _ = simulate_battery_discharge(batt, runs, optimal_params)

    # Save results
    save_results(batch_name, battery_number, cycle_number, saving_file_path,
                 optimal_params, batt, runs, final_sim_voltage, true_voltage, Mid_SOC, DOD)

    return qMax_opt, Ro_opt, wr_opt


def save_results(batch_name: str, battery_number: int, cycle_number: int,
                 saving_file_path: str, optimal_params: Dict[str, float],
                 batt, runs, simulated_voltage: np.ndarray, true_voltage: List[float],
                 Mid_SOC: float, DOD: float) -> None:
    """
    Save optimization results and generate plots.

    Args:
        batch_name: Name of the battery batch
        battery_number: Battery number identifier
        cycle_number: Cycle number identifier
        saving_file_path: Base path for saving files
        optimal_params: Dictionary of optimal parameters
        batt: Battery model object
        runs: Simulation run data
        simulated_voltage: Simulated voltage data
        true_voltage: True voltage measurements
        Mid_SOC: Middle state of charge
        DOD: Depth of discharge
    """
    # Create directory structure
    battery_path = os.path.join(saving_file_path, batch_name, f'Battery{battery_number}')
    BatteryParameterEstimator.create_folder(battery_path)

    # Save key parameters
    params_file = os.path.join(battery_path, 'key_parameters.xlsx')
    BatteryParameterEstimator.write_excel_parameters(
        params_file, optimal_params['qMax'], optimal_params['Ro'], optimal_params['wr']
    )

    # Save detailed simulation results
    results_file = os.path.join(battery_path, f'{cycle_number}.xlsx')

    # Re-run simulation to get all state data
    batt.parameters['x0'] = batt.initialize()  # Reset initial state
    optimization_params = optimal_params

    # Simulate again for detailed results
    simulated_voltage_full, _ = simulate_battery_discharge(batt, runs, optimization_params)

    # Generate and save plot
    create_comparison_plot(simulated_voltage, true_voltage, Mid_SOC, DOD,
                           battery_number, cycle_number, batch_name, saving_file_path)


def create_comparison_plot(simulated_voltage: np.ndarray, true_voltage: List[float],
                           Mid_SOC: float, DOD: float, battery_number: int,
                           cycle_number: int, batch_name: str, saving_file_path: str) -> None:
    """
    Create and save comparison plot between simulated and observed voltage.

    Args:
        simulated_voltage: Simulated voltage data
        true_voltage: True voltage measurements
        Mid_SOC: Middle state of charge
        DOD: Depth of discharge
        battery_number: Battery number identifier
        cycle_number: Cycle number identifier
        batch_name: Name of the battery batch
        saving_file_path: Base path for saving files
    """
    # Calculate SOC ranges for plotting
    start_SOC = Mid_SOC + DOD / 2
    end_SOC = Mid_SOC - DOD / 2

    sample_SOC = BatteryParameterEstimator.DEFAULT_SOC_SAMPLES
    SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(sample_SOC) / (sample_SOC - 1), new_min=end_SOC, new_max=start_SOC
    )

    # Prepare observed data
    observed_SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(len(true_voltage)) / (len(true_voltage) - 1),
        new_min=end_SOC, new_max=start_SOC
    )

    # Create plot
    plt.figure(figsize=(10, 6))

    # Plot simulated data
    simulated_SOC_range = BatteryParameterEstimator.rescale_array(
        np.arange(len(simulated_voltage)) / (len(simulated_voltage) - 1),
        new_min=0, new_max=1
    )
    plt.plot(simulated_SOC_range[:len(simulated_voltage)], simulated_voltage,
             'b-', label='Simulated', linewidth=2)

    # Plot observed data
    func_observed = interp1d(np.flipud(observed_SOC_range), true_voltage, kind='cubic')
    observed_voltage_range = func_observed(SOC_range)
    plt.plot(SOC_range, observed_voltage_range, 'r--', label='Observed', linewidth=2)

    # Formatting
    plt.gca().invert_xaxis()
    plt.xlabel('State of Charge (SOC)')
    plt.ylabel('Voltage (V)')
    plt.title(f'Battery {battery_number} - Cycle {cycle_number}')
    plt.legend()
    plt.grid(True, alpha=0.3)

    # Save plot
    figure_path = os.path.join(saving_file_path, batch_name, f'Battery{battery_number}', 'Figure')
    BatteryParameterEstimator.create_folder(figure_path)
    plt.savefig(os.path.join(figure_path, f'{cycle_number}.jpg'), dpi=300, bbox_inches='tight')
    plt.show()


# Legacy function aliases for backward compatibility
def battery_working_condition_match(batch_name: str) -> Tuple[float, float]:
    """Legacy function - use BatteryParameterEstimator.get_battery_working_condition instead."""
    return BatteryParameterEstimator.get_battery_working_condition(batch_name)


def create_folder(folder_path: str) -> None:
    """Legacy function - use BatteryParameterEstimator.create_folder instead."""
    BatteryParameterEstimator.create_folder(folder_path)


def write_excel_important_parameters(excel_path: str, qMax: float, R0: float, wr: float) -> None:
    """Legacy function - use BatteryParameterEstimator.write_excel_parameters instead."""
    BatteryParameterEstimator.write_excel_parameters(excel_path, qMax, R0, wr)


def write_excel_results(excel_path: str, values: List[float]) -> None:
    """Legacy function - use BatteryParameterEstimator.write_excel_results instead."""
    BatteryParameterEstimator.write_excel_results(excel_path, values)

